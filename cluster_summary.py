#!/usr/bin/env python3
"""
VMware Cluster Summary Generator - Excel Format
Produces a formatted Excel workbook with per-cluster tables.
Each cluster has: Summary section (key metrics) + Host details table.
Uses key-based joins (MOIDs/UUIDs) - does NOT rely on cluster/VM names.

Required: pip install pandas openpyxl
"""

import pandas as pd
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("WARNING: openpyxl not installed. Run: pip install openpyxl")
    print("Falling back to CSV output.")
    OPENPYXL_AVAILABLE = False

# ============================================================================
# CONFIGURATION
# ============================================================================
# Use the directory where the script is located
SCRIPT_DIR = Path(__file__).parent.resolve()
INPUT_FILES_DIR = SCRIPT_DIR / "input_files"  # Place collector Excel files here
RESULTS_DIR = SCRIPT_DIR / "results"
DATE_STR = datetime.now().strftime("%Y-%m-%d")

# Create folders if they don't exist
INPUT_FILES_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

# ============================================================================
# EXCEL EXTRACTION
# ============================================================================

def extract_excel_sheets(excel_path, output_dir):
    """
    Extract all sheets from a collector Excel file to individual CSVs.
    Returns True if required sheets were found, False otherwise.
    """
    print(f"  Extracting sheets from {excel_path.name}...")
    
    try:
        xl = pd.ExcelFile(excel_path)
        sheet_names = xl.sheet_names
        print(f"    Found {len(sheet_names)} sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
        
        # Required sheets for processing
        required = ['vCluster', 'vHosts', 'vInfo', 'vCPU', 'vMemory', 'vDisk', 'vPartition']
        
        extracted = []
        for sheet in sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet)
            csv_path = output_dir / f"{sheet}.csv"
            df.to_csv(csv_path, index=False)
            extracted.append(sheet)
        
        # Check for required sheets
        missing = [s for s in required if s not in extracted]
        if missing:
            print(f"    WARNING: Missing required sheets: {', '.join(missing)}")
            return False
        
        print(f"    Extracted {len(extracted)} sheets to CSVs")
        return True
        
    except Exception as e:
        print(f"    ERROR extracting Excel: {e}")
        return False

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def mib_to_tib(mib):
    """Convert MiB to TiB (binary)"""
    if pd.isna(mib):
        return 0
    return mib / (1024 * 1024)

def mib_to_gib(mib):
    """Convert MiB to GiB (binary)"""
    if pd.isna(mib):
        return 0
    return mib / 1024

def mib_to_mb(mib):
    """Convert MiB to MB (decimal, 1 MB = 1000^2 bytes)"""
    if pd.isna(mib):
        return 0
    return (mib * 1024 * 1024) / (1000 ** 2)

def mib_to_tb(mib):
    """Convert MiB to TB (decimal)"""
    if pd.isna(mib):
        return 0
    return (mib * 1024 * 1024) / (1000 ** 4)

def gb_to_gib(gb):
    """Convert GB (decimal) to GiB (binary)"""
    if pd.isna(gb):
        return 0
    return gb * (1000 ** 3) / (1024 ** 3)

def load_csv(filepath, name):
    """Load CSV with error handling and report row count"""
    try:
        df = pd.read_csv(filepath)
        df = df.replace('', pd.NA)
        return df
    except Exception as e:
        print(f"ERROR loading {name}: {e}")
        return pd.DataFrame()

def safe_round(val, decimals=2):
    """Safely round a value, returning 0 if not numeric"""
    try:
        return round(float(val), decimals)
    except:
        return 0

def normalize_column_names(df, column_mappings):
    """
    Normalize column names by renaming variations to expected names.
    
    Args:
        df: DataFrame to normalize
        column_mappings: Dict mapping expected_name -> list of possible variations
    
    Returns:
        DataFrame with renamed columns
    """
    rename_map = {}
    for expected_name, variations in column_mappings.items():
        if expected_name not in df.columns:
            for variation in variations:
                if variation in df.columns:
                    rename_map[variation] = expected_name
                    break
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

def get_column_safe(df, column_name, default=0):
    """
    Safely get a column from DataFrame, returning default if not present.
    
    Args:
        df: DataFrame to access
        column_name: Name of column to retrieve
        default: Default value if column doesn't exist
    
    Returns:
        Column Series or Series of default values
    """
    if column_name in df.columns:
        return df[column_name]
    return pd.Series([default] * len(df), index=df.index)

# Column name mappings for vHosts - maps expected name to possible variations
VHOSTS_COLUMN_MAPPINGS = {
    'Free Space (MiB)': ['Freespace (MiB)', 'FreeSpace (MiB)', 'Free Space(MiB)', 'Freespace(MiB)'],
    'Service Tag': ['ServiceTag', 'Service_Tag', 'Serial Number', 'Serial'],
    'Maintenance Mode': ['MaintenanceMode', 'Maintenance_Mode', 'Maint Mode'],
}

# ============================================================================
# FILE SAVING WITH ERROR HANDLING
# ============================================================================

def save_file_with_retry(save_func, filepath, max_retries=3):
    """
    Attempt to save a file with retry logic for permission errors.
    If file is locked, tries alternative filenames with timestamps.
    
    Args:
        save_func: Callable that takes filepath and saves the file
        filepath: Path object or string for the target file
        max_retries: Number of alternative filenames to try
    
    Returns:
        Path: The actual path where the file was saved
    
    Raises:
        PermissionError: If all retries fail
    """
    from datetime import datetime
    filepath = Path(filepath)
    
    # First attempt: try the original filename
    try:
        save_func(filepath)
        return filepath
    except PermissionError as e:
        print(f"  WARNING: Cannot save to {filepath.name} - file may be open in another application.")
    
    # Retry with alternative filenames
    for retry in range(1, max_retries + 1):
        timestamp = datetime.now().strftime("%H%M%S")
        alt_name = f"{filepath.stem}_{timestamp}{filepath.suffix}"
        alt_path = filepath.parent / alt_name
        
        try:
            print(f"  Trying alternative filename: {alt_name}")
            save_func(alt_path)
            print(f"  Successfully saved to: {alt_path}")
            return alt_path
        except PermissionError:
            if retry < max_retries:
                import time
                time.sleep(0.5)  # Brief pause before next retry
                continue
    
    # All retries failed
    raise PermissionError(
        f"Cannot save file - all attempts failed.\n"
        f"  Target: {filepath}\n"
        f"  Please close the file if it's open in Excel or another application,\n"
        f"  then run the script again."
    )

def save_text_file_safe(filepath, content):
    """
    Save text content to a file with error handling.
    
    Args:
        filepath: Path to save to
        content: String content to write
    
    Returns:
        Path: The actual path where the file was saved, or None if failed
    """
    def do_save(path):
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
    
    try:
        return save_file_with_retry(do_save, filepath)
    except PermissionError as e:
        print(f"  WARNING: Could not save validation report: {e}")
        return None

# ============================================================================
# EXCEL STYLING
# ============================================================================

def get_styles():
    """Return style definitions for Excel formatting"""
    if not OPENPYXL_AVAILABLE:
        return {}
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return {
        'cluster_title': Font(bold=True, size=14, color='FFFFFF'),
        'cluster_title_fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
        'summary_label': Font(bold=True, size=10),
        'summary_value': Font(size=10),
        'summary_fill': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
        'host_header': Font(bold=True, size=10, color='FFFFFF'),
        'host_header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'host_data': Font(size=9),
        'host_alt_fill': PatternFill(start_color='E8EBF0', end_color='E8EBF0', fill_type='solid'),
        'border': thin_border,
        'center': Alignment(horizontal='center', vertical='center'),
        'left': Alignment(horizontal='left', vertical='center'),
        'right': Alignment(horizontal='right', vertical='center'),
    }

# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process_collector(input_dir, output_prefix, output_excel, output_csv, validation_report):
    """
    Process a single collector's CSV files and generate outputs.
    
    Args:
        input_dir: Path to directory containing CSV files
        output_prefix: Prefix for output identification
        output_excel: Path TO output Excel file
        output_csv: Path to output CSV file (fallback)
        validation_report: Path to validation report
    """
    validation_lines = []
    validation_lines.append("=" * 80)
    validation_lines.append("VMware Cluster Summary - Validation Report")
    validation_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    validation_lines.append("=" * 80)
    validation_lines.append("")
    
    # -------------------------------------------------------------------------
    # 1. LOAD ALL CSVs
    # -------------------------------------------------------------------------
    print("Loading CSV files...")
    
    vcluster = load_csv(input_dir / "vCluster.csv", "vCluster")
    vhosts = load_csv(input_dir / "vHosts.csv", "vHosts")
    vinfo = load_csv(input_dir / "vInfo.csv", "vInfo")
    vcpu = load_csv(input_dir / "vCPU.csv", "vCPU")
    vmemory = load_csv(input_dir / "vMemory.csv", "vMemory")
    vdisk = load_csv(input_dir / "vDisk.csv", "vDisk")
    vpartition = load_csv(input_dir / "vPartition.csv", "vPartition")
    
    validation_lines.append("CSV Row Counts:")
    validation_lines.append("-" * 50)
    validation_lines.append(f"  vCluster.csv:    {len(vcluster):>6} rows")
    validation_lines.append(f"  vHosts.csv:      {len(vhosts):>6} rows")
    validation_lines.append(f"  vInfo.csv:       {len(vinfo):>6} rows")
    validation_lines.append(f"  vCPU.csv:        {len(vcpu):>6} rows")
    validation_lines.append(f"  vMemory.csv:     {len(vmemory):>6} rows")
    validation_lines.append(f"  vDisk.csv:       {len(vdisk):>6} rows")
    validation_lines.append(f"  vPartition.csv:  {len(vpartition):>6} rows")
    validation_lines.append("")
    
    # -------------------------------------------------------------------------
    # 2. BUILD HOST -> CLUSTER MAPPING
    # -------------------------------------------------------------------------
    print("Building host -> cluster mapping...")
    
    host_to_cluster = vhosts[['MOID', 'Cluster']].copy()
    host_to_cluster.columns = ['Host_MOID', 'Cluster_MOID']
    host_to_cluster = host_to_cluster.drop_duplicates(subset=['Host_MOID'])
    
    # -------------------------------------------------------------------------
    # 3. BUILD VM -> CLUSTER MAPPING
    # -------------------------------------------------------------------------
    print("Building VM -> cluster mapping...")
    
    vm_to_host = vinfo[['UUID', 'Host MOID']].copy()
    vm_to_host.columns = ['UUID', 'Host_MOID']
    vm_to_host = vm_to_host.drop_duplicates(subset=['UUID'])
    
    vm_to_cluster = vm_to_host.merge(
        host_to_cluster[['Host_MOID', 'Cluster_MOID']],
        on='Host_MOID',
        how='left'
    )
    
    vms_missing_cluster = vm_to_cluster['Cluster_MOID'].isna().sum()
    validation_lines.append(f"VMs in vInfo: {len(vm_to_cluster)}")
    validation_lines.append(f"VMs missing cluster mapping: {vms_missing_cluster}")
    validation_lines.append("")
    
    # -------------------------------------------------------------------------
    # 4. MAP VMs TO CLUSTERS
    # -------------------------------------------------------------------------
    print("Mapping VMs to clusters...")
    
    vcpu_with_cluster = vcpu.merge(vm_to_cluster[['UUID', 'Cluster_MOID']], on='UUID', how='left')
    vmem_with_cluster = vmemory.merge(vm_to_cluster[['UUID', 'Cluster_MOID']], on='UUID', how='left')
    vdisk_with_cluster = vdisk.merge(vm_to_cluster[['UUID', 'Cluster_MOID']], on='UUID', how='left')
    vpart_with_cluster = vpartition.merge(vm_to_cluster[['UUID', 'Cluster_MOID']], on='UUID', how='left')
    
    # -------------------------------------------------------------------------
    # 5. AGGREGATE BY CLUSTER
    # -------------------------------------------------------------------------
    print("Aggregating metrics by cluster...")
    
    clusters = vcluster[['Datacenter', 'MOID', 'Cluster Name']].copy()
    clusters.columns = ['Datacenter_MOID', 'Cluster_MOID', 'Cluster_Name']
    
    # VM counts
    vm_counts = vm_to_cluster.groupby('Cluster_MOID').size().reset_index(name='Total_VMs')
    clusters = clusters.merge(vm_counts, on='Cluster_MOID', how='left')
    clusters['Total_VMs'] = clusters['Total_VMs'].fillna(0).astype(int)
    
    # vCPUs
    vcpu_agg = vcpu_with_cluster[vcpu_with_cluster['Cluster_MOID'].notna()].copy()
    vcpu_agg['vCPUs'] = pd.to_numeric(vcpu_agg['vCPUs'], errors='coerce').fillna(0)
    vcpu_totals = vcpu_agg.groupby('Cluster_MOID')['vCPUs'].sum().reset_index(name='Total_vCPU')
    clusters = clusters.merge(vcpu_totals, on='Cluster_MOID', how='left')
    clusters['Total_vCPU'] = clusters['Total_vCPU'].fillna(0).astype(int)
    
    # Max vCPU per VM
    vcpu_per_vm = vcpu_agg.groupby(['Cluster_MOID', 'UUID'])['vCPUs'].max().reset_index()
    max_vcpu = vcpu_per_vm.groupby('Cluster_MOID')['vCPUs'].max().reset_index(name='Max_VM_vCPU')
    clusters = clusters.merge(max_vcpu, on='Cluster_MOID', how='left')
    clusters['Max_VM_vCPU'] = clusters['Max_VM_vCPU'].fillna(0).astype(int)
    
    # RAM
    vmem_agg = vmem_with_cluster[vmem_with_cluster['Cluster_MOID'].notna()].copy()
    vmem_agg['Size (MiB)'] = pd.to_numeric(vmem_agg['Size (MiB)'], errors='coerce').fillna(0)
    vmem_per_vm = vmem_agg.groupby(['Cluster_MOID', 'UUID'])['Size (MiB)'].max().reset_index()
    ram_totals = vmem_per_vm.groupby('Cluster_MOID')['Size (MiB)'].sum().reset_index(name='Total_RAM_MiB')
    clusters = clusters.merge(ram_totals, on='Cluster_MOID', how='left')
    clusters['Total_RAM_MiB'] = clusters['Total_RAM_MiB'].fillna(0)
    clusters['Total_RAM_GiB'] = clusters['Total_RAM_MiB'].apply(mib_to_gib)
    clusters['Total_RAM_TiB'] = clusters['Total_RAM_MiB'].apply(mib_to_tib)
    
    # Max RAM per VM
    max_ram = vmem_per_vm.groupby('Cluster_MOID')['Size (MiB)'].max().reset_index(name='Max_VM_RAM_MiB')
    clusters = clusters.merge(max_ram, on='Cluster_MOID', how='left')
    clusters['Max_VM_RAM_MiB'] = clusters['Max_VM_RAM_MiB'].fillna(0)
    clusters['Max_VM_RAM_GiB'] = clusters['Max_VM_RAM_MiB'].apply(mib_to_gib)
    
    # vDisk capacity
    vdisk_agg = vdisk_with_cluster[vdisk_with_cluster['Cluster_MOID'].notna()].copy()
    vdisk_agg['Capacity (MiB)'] = pd.to_numeric(vdisk_agg['Capacity (MiB)'], errors='coerce').fillna(0)
    vdisk_per_vm = vdisk_agg.groupby(['Cluster_MOID', 'UUID'])['Capacity (MiB)'].sum().reset_index(name='VM_Capacity_MiB')
    capacity_totals = vdisk_per_vm.groupby('Cluster_MOID')['VM_Capacity_MiB'].sum().reset_index(name='Capacity_Storage_MiB')
    clusters = clusters.merge(capacity_totals, on='Cluster_MOID', how='left')
    clusters['Capacity_Storage_MiB'] = clusters['Capacity_Storage_MiB'].fillna(0)
    clusters['Capacity_Storage_GiB'] = clusters['Capacity_Storage_MiB'].apply(mib_to_gib)
    clusters['Capacity_Storage_TiB'] = clusters['Capacity_Storage_MiB'].apply(mib_to_tib)
    
    # Max vDisk per VM
    max_cap = vdisk_per_vm.groupby('Cluster_MOID')['VM_Capacity_MiB'].max().reset_index(name='Max_VM_Capacity_MiB')
    clusters = clusters.merge(max_cap, on='Cluster_MOID', how='left')
    clusters['Max_VM_Capacity_MiB'] = clusters['Max_VM_Capacity_MiB'].fillna(0)
    clusters['Max_VM_Capacity_GiB'] = clusters['Max_VM_Capacity_MiB'].apply(mib_to_gib)
    clusters['Max_VM_Capacity_TiB'] = clusters['Max_VM_Capacity_MiB'].apply(mib_to_tib)
    
    # vPartition provisioned/consumed
    vpart_agg = vpart_with_cluster[vpart_with_cluster['Cluster_MOID'].notna()].copy()
    vpart_agg['Consumed (MiB)'] = pd.to_numeric(vpart_agg['Consumed (MiB)'], errors='coerce').fillna(0)
    vpart_agg['Capacity (MiB)'] = pd.to_numeric(vpart_agg['Capacity (MiB)'], errors='coerce').fillna(0)
    vpart_per_vm = vpart_agg.groupby(['Cluster_MOID', 'UUID']).agg({
        'Consumed (MiB)': 'sum', 'Capacity (MiB)': 'sum'
    }).reset_index()
    vpart_per_vm.columns = ['Cluster_MOID', 'UUID', 'VM_Consumed_MiB', 'VM_Provisioned_MiB']
    vpart_totals = vpart_per_vm.groupby('Cluster_MOID').agg({
        'VM_Consumed_MiB': 'sum', 'VM_Provisioned_MiB': 'sum'
    }).reset_index()
    vpart_totals.columns = ['Cluster_MOID', 'Consumed_Storage_MiB', 'Provisioned_Storage_MiB']
    clusters = clusters.merge(vpart_totals, on='Cluster_MOID', how='left')
    clusters['Consumed_Storage_MiB'] = clusters['Consumed_Storage_MiB'].fillna(0)
    clusters['Provisioned_Storage_MiB'] = clusters['Provisioned_Storage_MiB'].fillna(0)
    clusters['Consumed_Storage_GiB'] = clusters['Consumed_Storage_MiB'].apply(mib_to_gib)
    clusters['Consumed_Storage_TiB'] = clusters['Consumed_Storage_MiB'].apply(mib_to_tib)
    clusters['Provisioned_Storage_GiB'] = clusters['Provisioned_Storage_MiB'].apply(mib_to_gib)
    clusters['Provisioned_Storage_TiB'] = clusters['Provisioned_Storage_MiB'].apply(mib_to_tib)
    
    # Max consumed/provisioned per VM
    max_vpart = vpart_per_vm.groupby('Cluster_MOID').agg({
        'VM_Consumed_MiB': 'max', 'VM_Provisioned_MiB': 'max'
    }).reset_index()
    max_vpart.columns = ['Cluster_MOID', 'Max_VM_Consumed_MiB', 'Max_VM_Provisioned_MiB']
    clusters = clusters.merge(max_vpart, on='Cluster_MOID', how='left')
    clusters['Max_VM_Consumed_MiB'] = clusters['Max_VM_Consumed_MiB'].fillna(0)
    clusters['Max_VM_Provisioned_MiB'] = clusters['Max_VM_Provisioned_MiB'].fillna(0)
    clusters['Max_VM_Consumed_GiB'] = clusters['Max_VM_Consumed_MiB'].apply(mib_to_gib)
    clusters['Max_VM_Consumed_TiB'] = clusters['Max_VM_Consumed_MiB'].apply(mib_to_tib)
    clusters['Max_VM_Provisioned_GiB'] = clusters['Max_VM_Provisioned_MiB'].apply(mib_to_gib)
    clusters['Max_VM_Provisioned_TiB'] = clusters['Max_VM_Provisioned_MiB'].apply(mib_to_tib)
    
    # -------------------------------------------------------------------------
    # 6. HOST AGGREGATES FOR CLUSTER SUMMARY
    # -------------------------------------------------------------------------
    print("Calculating host aggregates...")
    
    # Prepare host data with numeric conversions
    host_df = vhosts.copy()
    
    # Normalize column names to handle variations across collector versions
    host_df = normalize_column_names(host_df, VHOSTS_COLUMN_MAPPINGS)
    
    host_df['CPUs'] = pd.to_numeric(host_df['CPUs'], errors='coerce').fillna(0).astype(int)
    host_df['CPU Cores'] = pd.to_numeric(host_df['CPU Cores'], errors='coerce').fillna(0).astype(int)
    host_df['Cores per CPU'] = pd.to_numeric(host_df['Cores per CPU'], errors='coerce').fillna(0).astype(int)
    host_df['Memory Size'] = pd.to_numeric(host_df['Memory Size'], errors='coerce').fillna(0)
    host_df['VMs'] = pd.to_numeric(host_df['VMs'], errors='coerce').fillna(0).astype(int)
    host_df['NICs'] = pd.to_numeric(host_df['NICs'], errors='coerce').fillna(0).astype(int)
    host_df['CPU Speed'] = pd.to_numeric(host_df['CPU Speed'], errors='coerce').fillna(0)
    host_df['Capacity (MiB)'] = pd.to_numeric(host_df['Capacity (MiB)'], errors='coerce').fillna(0)
    host_df['Consumed (MiB)'] = pd.to_numeric(host_df['Consumed (MiB)'], errors='coerce').fillna(0)
    host_df['Free Space (MiB)'] = pd.to_numeric(get_column_safe(host_df, 'Free Space (MiB)', 0), errors='coerce').fillna(0)
    host_df['GPU Count'] = pd.to_numeric(get_column_safe(host_df, 'GPU Count', 0), errors='coerce').fillna(0).astype(int)
    host_df['GPU Memory Size (MiB)'] = pd.to_numeric(get_column_safe(host_df, 'GPU Memory Size (MiB)', 0), errors='coerce').fillna(0)
    
    # Aggregate per cluster
    host_agg = host_df.groupby('Cluster').agg({
        'MOID': 'nunique',
        'CPUs': 'sum',
        'CPU Cores': 'sum',
        'Cores per CPU': 'first',
        'Memory Size': 'sum',
        'Capacity (MiB)': 'sum',
        'Consumed (MiB)': 'sum',
        'Free Space (MiB)': 'sum',
    }).reset_index()
    host_agg.columns = ['Cluster_MOID', 'No_Hosts', 'Total_CPUs', 'Total_pCores', 'Cores_per_CPU',
                        'Total_Host_RAM_GB', 'Host_Capacity_MiB', 'Host_Consumed_MiB', 'Host_Free_MiB']
    
    clusters = clusters.merge(host_agg, on='Cluster_MOID', how='left')
    clusters['No_Hosts'] = clusters['No_Hosts'].fillna(0).astype(int)
    clusters['Total_CPUs'] = clusters['Total_CPUs'].fillna(0).astype(int)
    clusters['Total_pCores'] = clusters['Total_pCores'].fillna(0).astype(int)
    clusters['Cores_per_CPU'] = clusters['Cores_per_CPU'].fillna(0).astype(int)
    clusters['Total_Host_RAM_GB'] = clusters['Total_Host_RAM_GB'].fillna(0)
    clusters['Total_Host_RAM_GiB'] = clusters['Total_Host_RAM_GB'].apply(gb_to_gib)
    clusters['Host_Capacity_MiB'] = clusters['Host_Capacity_MiB'].fillna(0)
    clusters['Host_Capacity_TiB'] = clusters['Host_Capacity_MiB'].apply(mib_to_tib)
    clusters['Host_Consumed_MiB'] = clusters['Host_Consumed_MiB'].fillna(0)
    clusters['Host_Consumed_TiB'] = clusters['Host_Consumed_MiB'].apply(mib_to_tib)
    clusters['Host_Free_MiB'] = clusters['Host_Free_MiB'].fillna(0)
    clusters['Host_Free_TiB'] = clusters['Host_Free_MiB'].apply(mib_to_tib)
    
    # Get distinct hardware models per cluster
    models_per_cluster = host_df.groupby('Cluster')['Model'].apply(
        lambda x: ' | '.join(sorted(x.dropna().unique()))
    ).reset_index()
    models_per_cluster.columns = ['Cluster_MOID', 'Host_Models']
    clusters = clusters.merge(models_per_cluster, on='Cluster_MOID', how='left')
    
    # Get distinct vendors per cluster
    vendors_per_cluster = host_df.groupby('Cluster')['Vendor'].apply(
        lambda x: ' | '.join(sorted(x.dropna().unique()))
    ).reset_index()
    vendors_per_cluster.columns = ['Cluster_MOID', 'Vendors']
    clusters = clusters.merge(vendors_per_cluster, on='Cluster_MOID', how='left')
    
    # Get distinct CPU models per cluster
    cpu_models_per_cluster = host_df.groupby('Cluster')['CPU Model'].apply(
        lambda x: ' | '.join(sorted(x.dropna().unique()))
    ).reset_index()
    cpu_models_per_cluster.columns = ['Cluster_MOID', 'CPU_Models']
    clusters = clusters.merge(cpu_models_per_cluster, on='Cluster_MOID', how='left')
    
    # vCPU:pCore ratio
    clusters['vCPU_pCore_Ratio'] = clusters.apply(
        lambda row: round(row['Total_vCPU'] / row['Total_pCores'], 2) if row['Total_pCores'] > 0 else 0,
        axis=1
    )
    
    # Avg vCPU per VM
    clusters['Avg_vCPU_per_VM'] = clusters.apply(
        lambda row: round(row['Total_vCPU'] / row['Total_VMs'], 2) if row['Total_VMs'] > 0 else 0,
        axis=1
    )
    
    # Sort clusters
    clusters = clusters.sort_values(['Datacenter_MOID', 'Cluster_MOID'])
    
    # -------------------------------------------------------------------------
    # 7. DEFINE HOST TABLE COLUMNS
    # -------------------------------------------------------------------------
    
    host_table_columns = [
        ('Host MOID', 'MOID'),
        ('Service Tag', 'Service Tag'),
        ('Vendor', 'Vendor'),
        ('Model', 'Model'),
        ('BIOS', 'BIOS'),
        ('Hypervisor', 'Hypervisor'),
        ('Maint Mode', 'Maintenance Mode'),
        ('CPU Sockets', 'CPUs'),
        ('CPU Model', 'CPU Model'),
        ('Total Cores', 'CPU Cores'),
        ('Cores/CPU', 'Cores per CPU'),
        ('CPU MHz', 'CPU Speed'),
        ('RAM (GB)', 'Memory Size'),
        ('NICs', 'NICs'),
        ('GPUs', 'GPU Count'),
        ('Storage Cap (TiB)', None),
        ('Storage Used (TiB)', None),
        ('Storage Free (TiB)', None),
        ('VMs', 'VMs'),
    ]
    
    # -------------------------------------------------------------------------
    # 8. BUILD EXCEL OUTPUT
    # -------------------------------------------------------------------------
    
    if OPENPYXL_AVAILABLE:
        print("Creating formatted Excel workbook...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cluster Summary"
        
        styles = get_styles()
        current_row = 1
        
        # ===== AGGREGATE SUMMARY TABLE (ALL CLUSTERS) =====
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        title_cell = ws.cell(row=current_row, column=1, value="TOTAL SUMMARY (ALL CLUSTERS)")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        title_cell.alignment = styles['left']
        current_row += 1
        
        # Aggregate totals
        total_hosts = clusters['No_Hosts'].sum()
        total_vms = clusters['Total_VMs'].sum()
        total_vcpu = clusters['Total_vCPU'].sum()
        total_pcores = clusters['Total_pCores'].sum()
        total_host_ram_gb = clusters['Total_Host_RAM_GB'].sum()
        total_vm_ram_gib = clusters['Total_RAM_GiB'].sum()
        total_vm_ram_tib = clusters['Total_RAM_TiB'].sum()
        total_provisioned_tib = clusters['Provisioned_Storage_TiB'].sum()
        total_consumed_tib = clusters['Consumed_Storage_TiB'].sum()
        total_capacity_tib = clusters['Capacity_Storage_TiB'].sum()
        total_cluster_capacity_tib = clusters['Host_Capacity_TiB'].sum()
        total_cluster_consumed_tib = clusters['Host_Consumed_TiB'].sum()
        total_cluster_free_tib = clusters['Host_Free_TiB'].sum()
        overall_vcpu_pcore = round(total_vcpu / total_pcores, 2) if total_pcores > 0 else 0
        overall_avg_vcpu_vm = round(total_vcpu / total_vms, 2) if total_vms > 0 else 0
        max_vm_vcpu = clusters['Max_VM_vCPU'].max()
        max_vm_ram_gib = clusters['Max_VM_RAM_GiB'].max()
        max_vm_provisioned_gib = clusters['Max_VM_Provisioned_GiB'].max()
        max_vm_capacity_gib = clusters['Max_VM_Capacity_GiB'].max()
        
        # All unique hardware
        all_vendors = ' | '.join(sorted(set(' | '.join(clusters['Vendors'].dropna()).split(' | '))))
        all_models = ' | '.join(sorted(set(' | '.join(clusters['Host_Models'].dropna()).split(' | '))))
        all_cpus = ' | '.join(sorted(set(' | '.join(clusters['CPU_Models'].dropna()).split(' | '))))
        
        aggregate_data = [
            ('Total Clusters', len(clusters)),
            ('Total Hosts', total_hosts),
            ('Total VMs', total_vms),
            ('Total vCPU', total_vcpu),
            ('Total pCores', total_pcores),
            ('vCPU:pCore Ratio', overall_vcpu_pcore),
            ('Avg vCPU/VM', overall_avg_vcpu_vm),
            ('Total Host RAM (GB)', safe_round(total_host_ram_gb, 2)),
            ('VM RAM Total (GiB)', safe_round(total_vm_ram_gib, 2)),
            ('VM RAM Total (TiB)', safe_round(total_vm_ram_tib, 4)),
            ('Max VM vCPU', max_vm_vcpu),
            ('Max VM RAM (GiB)', safe_round(max_vm_ram_gib, 2)),
            ('Provisioned Storage (TiB)', safe_round(total_provisioned_tib, 4)),
            ('Consumed Storage (TiB)', safe_round(total_consumed_tib, 4)),
            ('Capacity vDisk (TiB)', safe_round(total_capacity_tib, 4)),
            ('Max VM Provisioned (GiB)', safe_round(max_vm_provisioned_gib, 2)),
            ('Max VM Capacity (GiB)', safe_round(max_vm_capacity_gib, 2)),
            ('Cluster Storage Capacity (TiB)', safe_round(total_cluster_capacity_tib, 4)),
            ('Cluster Storage Consumed (TiB)', safe_round(total_cluster_consumed_tib, 4)),
            ('Cluster Storage Free (TiB)', safe_round(total_cluster_free_tib, 4)),
            ('Hardware Vendors', all_vendors),
            ('Hardware Models', all_models),
            ('CPU Models', all_cpus),
        ]
        
        # Write aggregate summary in 4-column layout
        agg_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        for i, (label, value) in enumerate(aggregate_data):
            row_offset = i // 2
            col_offset = (i % 2) * 2
            
            label_cell = ws.cell(row=current_row + row_offset, column=1 + col_offset, value=label)
            label_cell.font = styles['summary_label']
            label_cell.fill = agg_fill
            label_cell.border = styles['border']
            
            value_cell = ws.cell(row=current_row + row_offset, column=2 + col_offset, value=value)
            value_cell.font = styles['summary_value']
            value_cell.border = styles['border']
        
        current_row += (len(aggregate_data) + 1) // 2 + 2  # Extra blank row after summary
        
        # ===== PER-CLUSTER SECTIONS =====
        for idx, cluster in clusters.iterrows():
            cluster_moid = cluster['Cluster_MOID']
            cluster_name = cluster['Cluster_Name']
            
            # ===== CLUSTER TITLE =====
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            title_cell = ws.cell(row=current_row, column=1, 
                                 value=f"CLUSTER: {cluster_name} ({cluster_moid})")
            title_cell.font = styles['cluster_title']
            title_cell.fill = styles['cluster_title_fill']
            title_cell.alignment = styles['left']
            current_row += 1
            
            # ===== CLUSTER SUMMARY SECTION =====
            summary_data = [
                ('Datacenter MOID', cluster['Datacenter_MOID']),
                ('Cluster MOID', cluster_moid),
                ('Total Hosts', cluster['No_Hosts']),
                ('Total VMs', cluster['Total_VMs']),
                ('Total vCPU', cluster['Total_vCPU']),
                ('Total pCores', cluster['Total_pCores']),
                ('vCPU:pCore Ratio', cluster['vCPU_pCore_Ratio']),
                ('Avg vCPU/VM', cluster['Avg_vCPU_per_VM']),
                ('Total Host RAM (GB)', safe_round(cluster['Total_Host_RAM_GB'], 2)),
                ('VM RAM Total (GiB)', safe_round(cluster['Total_RAM_GiB'], 2)),
                ('VM RAM Total (TiB)', safe_round(cluster['Total_RAM_TiB'], 4)),
                ('Max VM vCPU', cluster['Max_VM_vCPU']),
                ('Max VM RAM (GiB)', safe_round(cluster['Max_VM_RAM_GiB'], 2)),
                ('Provisioned Storage (TiB)', safe_round(cluster['Provisioned_Storage_TiB'], 4)),
                ('Consumed Storage (TiB)', safe_round(cluster['Consumed_Storage_TiB'], 4)),
                ('Capacity vDisk (TiB)', safe_round(cluster['Capacity_Storage_TiB'], 4)),
                ('Max VM Provisioned (GiB)', safe_round(cluster['Max_VM_Provisioned_GiB'], 2)),
                ('Max VM Capacity (GiB)', safe_round(cluster['Max_VM_Capacity_GiB'], 2)),
                ('Cluster Storage Capacity (TiB)', safe_round(cluster['Host_Capacity_TiB'], 4)),
                ('Cluster Storage Consumed (TiB)', safe_round(cluster['Host_Consumed_TiB'], 4)),
                ('Cluster Storage Free (TiB)', safe_round(cluster['Host_Free_TiB'], 4)),
                ('Hardware Models', cluster.get('Host_Models', '')),
                ('Vendors', cluster.get('Vendors', '')),
                ('CPU Models', cluster.get('CPU_Models', '')),
            ]
            
            # Write summary in 4-column layout (Label, Value, Label, Value)
            for i, (label, value) in enumerate(summary_data):
                row_offset = i // 2
                col_offset = (i % 2) * 2
                
                label_cell = ws.cell(row=current_row + row_offset, column=1 + col_offset, value=label)
                label_cell.font = styles['summary_label']
                label_cell.fill = styles['summary_fill']
                label_cell.border = styles['border']
                
                value_cell = ws.cell(row=current_row + row_offset, column=2 + col_offset, value=value)
                value_cell.font = styles['summary_value']
                value_cell.border = styles['border']
            
            current_row += (len(summary_data) + 1) // 2 + 1
            
            # ===== HOST DETAILS TABLE =====
            ws.cell(row=current_row, column=1, value="HOST DETAILS").font = Font(bold=True, size=11)
            current_row += 1
            
            # Header row
            for col_idx, (header, _) in enumerate(host_table_columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = styles['host_header']
                cell.fill = styles['host_header_fill']
                cell.border = styles['border']
                cell.alignment = styles['center']
            current_row += 1
            
            # Host data rows
            cluster_hosts = host_df[host_df['Cluster'] == cluster_moid].copy()
            cluster_hosts = cluster_hosts.sort_values('MOID')
            
            for row_idx, (_, host) in enumerate(cluster_hosts.iterrows()):
                for col_idx, (header, source_col) in enumerate(host_table_columns, 1):
                    if source_col:
                        value = host.get(source_col, '')
                        if pd.isna(value):
                            value = ''
                    else:
                        # Calculated fields
                        if header == 'Storage Cap (TiB)':
                            value = safe_round(mib_to_tib(host['Capacity (MiB)']), 4)
                        elif header == 'Storage Used (TiB)':
                            value = safe_round(mib_to_tib(host['Consumed (MiB)']), 4)
                        elif header == 'Storage Free (TiB)':
                            value = safe_round(mib_to_tib(host['Free Space (MiB)']), 4)
                        else:
                            value = ''
                    
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.font = styles['host_data']
                    cell.border = styles['border']
                    
                    # Alternate row coloring
                    if row_idx % 2 == 1:
                        cell.fill = styles['host_alt_fill']
                
                current_row += 1
            
            # Add blank rows between clusters
            current_row += 2
        
        # Auto-size columns
        for col_idx in range(1, len(host_table_columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
        
        # Make first few columns wider for text content
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['I'].width = 45  # CPU Model
        
        print(f"Saving Excel file to {output_excel}...")
        
        def save_workbook(path):
            wb.save(path)
        
        try:
            actual_path = save_file_with_retry(save_workbook, output_excel)
            print(f"Output: {actual_path}")
        except PermissionError as e:
            print(f"ERROR: {e}")
            raise
    
    else:
        # Fallback to CSV if openpyxl not available
        print("openpyxl not available - creating simplified CSV output...")
        
        all_columns = ['Row_Type', 'Cluster_Name', 'Cluster_MOID', 'Datacenter_MOID',
                       'Host_MOID', 'Service_Tag', 'Vendor', 'Model', 'CPU_Sockets',
                       'CPU_Model', 'CPU_Cores', 'RAM_GB', 'VMs', 'Total_Hosts',
                       'Total_VMs', 'Total_vCPU', 'Total_pCores', 'vCPU_pCore_Ratio']
        
        output_rows = []
        for idx, cluster in clusters.iterrows():
            cluster_moid = cluster['Cluster_MOID']
            
            # Header row
            output_rows.append({col: col for col in all_columns})
            
            # Cluster summary row
            output_rows.append({
                'Row_Type': 'CLUSTER_SUMMARY',
                'Cluster_Name': cluster['Cluster_Name'],
                'Cluster_MOID': cluster_moid,
                'Datacenter_MOID': cluster['Datacenter_MOID'],
                'Total_Hosts': cluster['No_Hosts'],
                'Total_VMs': cluster['Total_VMs'],
                'Total_vCPU': cluster['Total_vCPU'],
                'Total_pCores': cluster['Total_pCores'],
                'vCPU_pCore_Ratio': cluster['vCPU_pCore_Ratio'],
            })
            
            # Host rows
            cluster_hosts = host_df[host_df['Cluster'] == cluster_moid]
            for _, host in cluster_hosts.iterrows():
                output_rows.append({
                    'Row_Type': 'HOST_DETAIL',
                    'Cluster_Name': cluster['Cluster_Name'],
                    'Cluster_MOID': cluster_moid,
                    'Host_MOID': host['MOID'],
                    'Service_Tag': host.get('Service Tag', ''),
                    'Vendor': host.get('Vendor', ''),
                    'Model': host.get('Model', ''),
                    'CPU_Sockets': host['CPUs'],
                    'CPU_Model': host.get('CPU Model', ''),
                    'CPU_Cores': host['CPU Cores'],
                    'RAM_GB': safe_round(host['Memory Size'], 2),
                    'VMs': host['VMs'],
                })
            
            # Blank separator
            output_rows.append({col: '' for col in all_columns})
        
        def save_csv(path):
            pd.DataFrame(output_rows).to_csv(path, index=False, header=False)
        
        try:
            actual_path = save_file_with_retry(save_csv, output_csv)
            print(f"Output: {actual_path}")
        except PermissionError as e:
            print(f"ERROR: {e}")
            raise
    
    # -------------------------------------------------------------------------
    # 9. VALIDATION REPORT
    # -------------------------------------------------------------------------
    validation_lines.append("=" * 80)
    validation_lines.append("SUMMARY")
    validation_lines.append("=" * 80)
    validation_lines.append(f"Total clusters: {len(clusters)}")
    validation_lines.append(f"Total hosts: {len(vhosts)}")
    validation_lines.append(f"Total VMs: {clusters['Total_VMs'].sum()}")
    validation_lines.append(f"Total vCPUs: {clusters['Total_vCPU'].sum()}")
    validation_lines.append(f"Total Host RAM (GB): {clusters['Total_Host_RAM_GB'].sum():.2f}")
    validation_lines.append("")
    
    validation_lines.append("=" * 80)
    validation_lines.append("HARDWARE BY CLUSTER")
    validation_lines.append("=" * 80)
    for _, row in clusters.iterrows():
        validation_lines.append(f"\n{row['Cluster_Name']} ({row['Cluster_MOID']}):")
        validation_lines.append(f"  Vendors: {row.get('Vendors', 'N/A')}")
        validation_lines.append(f"  Models: {row.get('Host_Models', 'N/A')}")
    
    validation_content = '\n'.join(validation_lines)
    actual_validation_path = save_text_file_safe(validation_report, validation_content)
    if actual_validation_path:
        print(f"Validation: {actual_validation_path}")
    
    print("Done processing this collector!")
    return True


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """
    Main entry point. Processes:
    1. Excel files from input_files/ folder (extracts sheets to CSVs, then processes)
    2. CSV files in script directory (legacy mode for backward compatibility)
    """
    print("=" * 70)
    print("VMware Cluster Summary Generator")
    print(f"Date: {DATE_STR}")
    print("=" * 70)
    
    processed_count = 0
    
    # Check for Excel collector files in input_files/
    excel_files = list(INPUT_FILES_DIR.glob("*.xlsx")) + list(INPUT_FILES_DIR.glob("*.xls"))
    
    if excel_files:
        print(f"\nFound {len(excel_files)} collector file(s) in input_files/:")
        for f in excel_files:
            print(f"  - {f.name}")
        print()
        
        for excel_file in excel_files:
            print("=" * 70)
            print(f"Processing: {excel_file.name}")
            print("=" * 70)
            
            # Create temp directory for extracted CSVs
            collector_name = excel_file.stem  # filename without extension
            temp_csv_dir = RESULTS_DIR / f"_temp_{collector_name}"
            temp_csv_dir.mkdir(exist_ok=True)
            
            # Extract sheets to CSVs
            if extract_excel_sheets(excel_file, temp_csv_dir):
                # Generate output paths for this collector
                output_excel = RESULTS_DIR / f"cluster_summary_{collector_name}_{DATE_STR}.xlsx"
                output_csv = RESULTS_DIR / f"cluster_summary_{collector_name}_{DATE_STR}.csv"
                validation_report = RESULTS_DIR / f"validation_{collector_name}_{DATE_STR}.txt"
                
                # Process the extracted CSVs
                try:
                    process_collector(
                        input_dir=temp_csv_dir,
                        output_prefix=collector_name,
                        output_excel=output_excel,
                        output_csv=output_csv,
                        validation_report=validation_report
                    )
                    processed_count += 1
                except Exception as e:
                    print(f"ERROR processing {excel_file.name}: {e}")
                    import traceback
                    traceback.print_exc()
            
            # Clean up temp CSV directory
            try:
                import shutil
                shutil.rmtree(temp_csv_dir)
                print(f"  Cleaned up temp files")
            except:
                pass
            
            print()
    
    else:
        # Legacy mode: Check for CSVs in script directory
        print("\nNo Excel files in input_files/. Checking for CSVs in script directory...")
        
        if (SCRIPT_DIR / "vCluster.csv").exists():
            print("Found CSV files in script directory (legacy mode)")
            
            output_excel = RESULTS_DIR / f"cluster_summary_{DATE_STR}.xlsx"
            output_csv = RESULTS_DIR / f"cluster_summary_{DATE_STR}.csv"
            validation_report = RESULTS_DIR / f"validation_report_{DATE_STR}.txt"
            
            process_collector(
                input_dir=SCRIPT_DIR,
                output_prefix="local",
                output_excel=output_excel,
                output_csv=output_csv,
                validation_report=validation_report
            )
            processed_count += 1
        else:
            print("\nNo input files found!")
            print("Options:")
            print("  1. Place collector Excel files (.xlsx) in: input_files/")
            print("  2. Place CSV files (vCluster.csv, vHosts.csv, etc.) in script directory")
    
    print("\n" + "=" * 70)
    print(f"COMPLETE: Processed {processed_count} collector file(s)")
    print(f"Results saved to: {RESULTS_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
